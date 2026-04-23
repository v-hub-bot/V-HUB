import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

providers = [
{"business_name": "#1 A Docs Painting with Excellence", "phone": "352-753-0056"},
{"business_name": "24 Hour Cart Club", "phone": "352-330-1911"},
{"business_name": "4Wheels Mobile Golf Cart Wash", "phone": "817-368-3977"},
{"business_name": "A Cut Above Tile & Hardwood Installations Inc.", "phone": "352-433-8032"},
{"business_name": "A Good Neighbor Pest & Pool", "phone": "813-579-0699"},
{"business_name": "A-1 Devon Irrigation", "phone": "352-551-3478"},
{"business_name": "AB Landscaping & Tree Services LLC", "phone": "352-587-1323"},
{"business_name": "Abundant Life Organizing", "phone": "314-266-8764"},
{"business_name": "Advanced Pool & Spa Sales", "phone": "352-205-4270"},
{"business_name": "Aggato's Landscaping", "phone": "352-426-6476"},
{"business_name": "Alchemy by Alexx", "phone": ""},
{"business_name": "All About Carts", "phone": "352-409-2702"},
{"business_name": "All Air & Heat Inc.", "phone": "352-728-4554"},
{"business_name": "All American Muscle Moving LLC", "phone": "352-233-8494"},
{"business_name": "All Surface Rejuvenations Plus", "phone": "352-502-0707"},
{"business_name": "Alternative Home Health Care — Medical Transport", "phone": "352-352-1414"},
{"business_name": "Amedisys Home Health", "phone": "352-391-1416"},
{"business_name": "Amour Mobile Nails", "phone": "321-320-9044"},
{"business_name": "Angela Perez Landscaping", "phone": "352-470-2510"},
{"business_name": "Appliance Repair Specialists — The Villages", "phone": "352-461-8598"},
{"business_name": "Arianna's Perfect Clean", "phone": "479-434-8892"},
{"business_name": "Attorney Fisher — The Villages", "phone": "352-503-4111"},
{"business_name": "Autos On The Go — The Villages", "phone": "866-924-7447"},
{"business_name": "AvA Nails", "phone": "352-626-5000"},
{"business_name": "BB Nail Bar", "phone": "352-461-0184"},
{"business_name": "BC Kustom Karts", "phone": "352-561-8440"},
{"business_name": "BCN Law Firm — The Villages", "phone": "352-775-4739"},
{"business_name": "BEST Battery — Golf Cart Accessories & Lighting", "phone": "352-391-5582"},
{"business_name": "Baker's Golf Cars", "phone": "352-793-1680"},
{"business_name": "Barefoot Landscape & Design LLC", "phone": "352-461-9172"},
{"business_name": "Bark Busters — The Villages", "phone": "877-500-2275"},
{"business_name": "Battery Boys", "phone": "352-643-1241"},
{"business_name": "Battery Brothers", "phone": "352-205-9859"},
{"business_name": "Bearded Brothers Flooring", "phone": "352-239-7269"},
{"business_name": "Being Your Eyes Home Watch", "phone": "583-506-3215"},
{"business_name": "Bereket Moving and Storage", "phone": "813-345-2522"},
{"business_name": "Bertram Pest Solutions", "phone": "352-608-5010"},
{"business_name": "Better Lawns & Landscapes", "phone": "352-638-7108"},
{"business_name": "Betty Brow Co., The", "phone": "352-461-4752"},
{"business_name": "Blue Line Transportation Services", "phone": "352-844-1746"},
{"business_name": "Bob's Appliance Repair", "phone": "352-897-1002"},
{"business_name": "Brothers & Sons Landscaping", "phone": "352-406-9307"},
{"business_name": "Bryan's Landscaping & Handyman Service", "phone": "352-298-8073"},
{"business_name": "Buffalo's Barber Shop", "phone": "352-205-7693"},
{"business_name": "C&H Trees & Landscaping", "phone": "352-205-0475"},
{"business_name": "C&J's Mobile Auto Repair & Towing", "phone": "352-408-9357"},
{"business_name": "Cal's Barber & Beauty — Brownwood", "phone": "352-750-5334"},
{"business_name": "Cal's Barber & Beauty — Colony", "phone": "352-674-9308"},
{"business_name": "Cal's Barber & Beauty — Lake Sumter", "phone": "352-259-9100"},
{"business_name": "Cal's Barber & Beauty — Spanish Springs", "phone": "352-391-5842"},
{"business_name": "Campbells Canine Training", "phone": "352-460-4887"},
{"business_name": "Carr's Floor & Home Carpet One", "phone": "352-915-2484"},
{"business_name": "Cart To Cars Mobile Detailing", "phone": "636-696-6633"},
{"business_name": "CartFixer Mobile Golf Cart Repair", "phone": "352-433-5411"},
{"business_name": "Cat the Dog Trainer", "phone": "951-514-7347"},
{"business_name": "Central FL Screens & Outdoor Services LLC", "phone": "352-298-8633"},
{"business_name": "Chessari Sod & Landscaping", "phone": "352-209-3559"},
{"business_name": "Chloes Beautique LLC", "phone": "352-812-1808"},
{"business_name": "Christian Brothers Power Washing & Windows", "phone": "813-337-8985"},
{"business_name": "Clay Oaks Mobile Veterinary & Grooming", "phone": "352-817-9373"},
{"business_name": "Clay's Carts", "phone": "352-302-5559"},
{"business_name": "Comfort Keepers Home Care", "phone": "352-720-8870"},
{"business_name": "Computer Central FL", "phone": "352-314-3855"},
{"business_name": "Conroy's Roofing & Tree Trimming", "phone": "214-794-2623"},
{"business_name": "Coopers Pooper Scooper", "phone": "352-308-4851"},
{"business_name": "Craine-N-Sons", "phone": "352-204-6181"},
{"business_name": "Craine-N-Sons Moving Services", "phone": "352-460-2449"},
{"business_name": "D&D Golf Cart Rentals LLC", "phone": "352-418-7661"},
{"business_name": "D&M Appliance Repair LLC", "phone": "352-259-4200"},
{"business_name": "DPS Flooring Kitchen & Bath Inc.", "phone": "813-469-4298"},
{"business_name": "DeSantis AC & Appliances", "phone": "352-330-4433"},
{"business_name": "Diamond G AC & Heating", "phone": "813-263-2967"},
{"business_name": "Diana Caruso LMT — Mobile Massage", "phone": "941-451-9122"},
{"business_name": "Dixie Safe & Lock Services LLC", "phone": "352-259-4511"},
{"business_name": "Dog Training Spot LLC, The", "phone": "321-754-1457"},
{"business_name": "DogSmith — The Villages", "phone": "352-391-5060"},
{"business_name": "Donnie Gates Home Repairs", "phone": "352-801-2888"},
{"business_name": "Doorstep Detail Inc.", "phone": "857-318-5629"},
{"business_name": "Driveway Coatings of Wildwood", "phone": "352-250-2564"},
{"business_name": "Duncan Family Movers LLC", "phone": "352-364-7312"},
{"business_name": "EC Car Transport — The Villages", "phone": "800-438-9089"},
{"business_name": "EW Excavation Pool LLC", "phone": "407-446-6666"},
{"business_name": "El Matador Tile & Remodeling", "phone": "352-504-6011"},
{"business_name": "El Sol Hardscaping", "phone": "352-406-0785"},
{"business_name": "Elite Flooring of Central Florida", "phone": "352-552-2898"},
{"business_name": "Elise's Mobile Notary", "phone": "631-428-8114"},
{"business_name": "Experts Garage Door Service", "phone": "352-410-0566"},
{"business_name": "Exterior Experts LLC", "phone": "352-531-5117"},
{"business_name": "FLI Lawn Tune-Up", "phone": "352-904-7827"},
{"business_name": "Face-Time Aesthetics", "phone": "352-460-4268"},
{"business_name": "Faded Lounge Barber Shop", "phone": "352-261-7737"},
{"business_name": "Fairway Golf Car Mobile Services", "phone": "352-978-5208"},
{"business_name": "Family's Choice Electrical Repair", "phone": "352-751-3794"},
{"business_name": "Feel Safe Home Watch, LLC", "phone": "352-750-9360"},
{"business_name": "Fernview Farm Lawn Fertilization", "phone": "352-245-7905"},
{"business_name": "Fertigator Lawn Care", "phone": "352-314-2867"},
{"business_name": "Finn's Maintenance", "phone": "352-454-8222"},
{"business_name": "Finney's Professional Services", "phone": "352-753-3629"},
{"business_name": "First Class Junk Removing LLC", "phone": "321-441-0430"},
{"business_name": "Five Star Golf Cart Service LLC", "phone": "352-301-3706"},
{"business_name": "Floors of Distinction", "phone": "352-751-4389"},
{"business_name": "Freeman Landscaping", "phone": "352-454-4343"},
{"business_name": "From The Ground Up Landscaping & Tree Removal", "phone": "352-805-3721"},
{"business_name": "Fur and Away Mobile Pet Grooming", "phone": "352-849-2999"},
{"business_name": "Galaxy Home Solutions", "phone": "352-748-4868"},
{"business_name": "Garage Door Doctor, The", "phone": "352-257-3321"},
{"business_name": "Garage Doors The Villages", "phone": "352-290-8986"},
{"business_name": "Garrett's Mobile Golf Car Service", "phone": "352-454-2272"},
{"business_name": "Gene Green Tree Service", "phone": "352-619-6155"},
{"business_name": "Genesis Health Clubs — The Villages", "phone": "352-753-6910"},
{"business_name": "Good Steward Home Solutions", "phone": "352-399-0073"},
{"business_name": "Goodell Electric LLC", "phone": "727-810-7194"},
{"business_name": "Green Mountain Landscaping & Tree Services", "phone": "352-776-8343"},
{"business_name": "Groome Transportation — The Villages", "phone": "352-539-9664"},
{"business_name": "H&H Handyman Services", "phone": "352-399-0987"},
{"business_name": "Happy Dog Pet Sitting LLC", "phone": "352-561-2464"},
{"business_name": "Happy Home Pet Care", "phone": "352-652-9406"},
{"business_name": "Healing Hands Wellness for Life", "phone": "352-750-9069"},
{"business_name": "Hector Damas Landscaping", "phone": "352-793-0049"},
{"business_name": "Helen's Nails & Spa", "phone": "352-689-0100"},
{"business_name": "Helping Hand Moving Services", "phone": "352-633-0018"},
{"business_name": "Hema Rupnarain CPA PA", "phone": "352-351-9880"},
{"business_name": "Hilton Automotive Inc.", "phone": "352-854-6868"},
{"business_name": "Holistic Business Consultants", "phone": "352-391-2163"},
{"business_name": "HomeWell Care Services — The Villages", "phone": "352-504-3400"},
{"business_name": "House Calls Mobile Pet Clinic", "phone": "352-231-4343"},
{"business_name": "Iron Shield Heating & Air", "phone": "352-247-8923"},
{"business_name": "Iron Village Gym — Personal Training, The", "phone": "352-461-0722"},
{"business_name": "Irrigation Nation", "phone": "352-633-0097"},
{"business_name": "J&H Property and Land Solutions", "phone": "352-455-4365"},
{"business_name": "J&P Home Inspections LLC", "phone": "603-978-9074"},
{"business_name": "J&R's Landscaping", "phone": "352-267-8019"},
{"business_name": "J.A. Pavers LLC", "phone": "352-461-1148"},
{"business_name": "J.R. Pruning", "phone": "352-470-8598"},
{"business_name": "JB Golf Carts — Lighting & Custom Upgrades", "phone": "352-830-2882"},
{"business_name": "JJ Electricians LLC", "phone": "352-268-9707"},
{"business_name": "JMW Total Services LLC", "phone": "321-299-2881"},
{"business_name": "Jackson Hewitt Tax Service — The Villages", "phone": "352-369-0244"},
{"business_name": "Jeff's Appliance Service", "phone": "502-909-6518"},
{"business_name": "JoanieFit — Personal Training", "phone": "302-743-1891"},
{"business_name": "John's Landscaping", "phone": "352-259-5861"},
{"business_name": "Johnson Brothers Plumbing", "phone": "352-748-0200"},
{"business_name": "Jonathan's Appliance Services", "phone": "352-568-5629"},
{"business_name": "Jose's Trimming, Planting & Lawn", "phone": "954-600-1174"},
{"business_name": "K&W Flooring Pros Inc.", "phone": "352-430-5808"},
{"business_name": "KB Landscape Supply Inc.", "phone": "352-748-7625"},
{"business_name": "KK's Auto Detailing", "phone": "352-429-8991"},
{"business_name": "Karen of Sunshine Home Cleaning", "phone": "352-216-0484"},
{"business_name": "Karmen Madden - Weed & Lawn Care", "phone": "352-261-9825"},
{"business_name": "Kent Services — Mobile Mechanic", "phone": "352-461-2282"},
{"business_name": "Kings PowerWash Service LLC", "phone": "352-362-1326"},
{"business_name": "Kling Towing & Vehicle Transport", "phone": "352-728-8010"},
{"business_name": "L&L Cleaning Services", "phone": "352-561-4388"},
{"business_name": "La Vie Nails & Spa", "phone": "352-633-5239"},
{"business_name": "Lake County Electrician LLC", "phone": "352-406-6030"},
{"business_name": "Lake-Sumter Electric", "phone": "352-793-8092"},
{"business_name": "Later Gator Moving LLC", "phone": "352-275-1769"},
{"business_name": "Lavish Mobile Detailing", "phone": "352-284-4441"},
{"business_name": "Leelandscape LLC", "phone": "352-633-0542"},
{"business_name": "Legacy Clinic of Chiropractic & Massage", "phone": "352-259-0024"},
{"business_name": "Legacy Flooring & Paint LLC", "phone": "352-530-4009"},
{"business_name": "Level Up Epoxy and Painting LLC", "phone": "352-807-2503"},
{"business_name": "Liberty Moves Orlando", "phone": "407-729-9108"},
{"business_name": "Lifestyle Landscape LLC", "phone": "352-516-0516"},
{"business_name": "Local Mobile Mechanic — The Villages", "phone": "844-669-6324"},
{"business_name": "Lotus Skin Studio, The", "phone": "352-350-3747"},
{"business_name": "Lunar Nails", "phone": ""},
{"business_name": "Luxe Nails & Spa", "phone": "352-750-2634"},
{"business_name": "Luxury Tile Construction Services", "phone": "407-460-8604"},
{"business_name": "M&S Air Conditioning", "phone": "352-235-9178"},
{"business_name": "Make Room with Renée", "phone": "321-328-5231"},
{"business_name": "Marketplace Movers & Painters", "phone": "352-480-1286"},
{"business_name": "Marshall Properties & Remodeling LLC", "phone": "352-267-2481"},
{"business_name": "Massey Services — The Villages", "phone": "352-259-0500"},
{"business_name": "Mazza Innovations LLC", "phone": "352-399-1102"},
{"business_name": "McWilliams Painting", "phone": "352-530-0189"},
{"business_name": "Meryl's Pest Control Service", "phone": "352-694-1977"},
{"business_name": "Michael's Termite & Pest Control", "phone": "352-425-4783"},
{"business_name": "Mike Scott Plumbing — The Villages", "phone": "352-748-9111"},
{"business_name": "Miller & Sons Plumbing", "phone": "352-748-5533"},
{"business_name": "Millhorn & Shanawany Law Firm", "phone": "352-205-4995"},
{"business_name": "Miracle Cleaning Services", "phone": "352-216-4801"},
{"business_name": "Miso Nails & Head Spa", "phone": "352-399-2561"},
{"business_name": "Mobile Bridal Beauty LLC", "phone": "904-252-2143"},
{"business_name": "Morgan's Outside Service", "phone": "352-454-6218"},
{"business_name": "Mr. Fix It Garage Doors", "phone": "877-561-8443"},
{"business_name": "Mr. Goodbrush Painting & Renovations", "phone": "407-443-6032"},
{"business_name": "Mr. Greg Enterprises Inc.", "phone": "352-426-7162"},
{"business_name": "Mr. Handyman", "phone": "352-259-6566"},
{"business_name": "Nature's Garden", "phone": "352-504-1345"},
{"business_name": "Nails & Beyond", "phone": "352-753-9222"},
{"business_name": "Nails & Beyond 2", "phone": "352-751-6777"},
{"business_name": "Nikky Pets Mobile Grooming", "phone": "954-336-6888"},
{"business_name": "NorthStar Business Consultants", "phone": "800-966-3681"},
{"business_name": "Ocala Fence LLC", "phone": "352-274-0823"},
{"business_name": "Ollie's Auto and Tire", "phone": "352-561-4307"},
{"business_name": "One Hour Heating & Air Conditioning", "phone": "352-897-9700"},
{"business_name": "Oscar's Quality Painting LLC", "phone": "352-361-4414"},
{"business_name": "P&J Flawless Paws Mobile Grooming", "phone": "352-561-5900"},
{"business_name": "PC Wizards — The Villages", "phone": "352-350-0610"},
{"business_name": "PLOS Lawn & Landscaping", "phone": "352-454-9027"},
{"business_name": "PM by Dawn — Permanent Makeup Boutique", "phone": "352-775-1551"},
{"business_name": "Paws Animal Hospital — Home Vet Calls", "phone": "352-259-2439"},
{"business_name": "Pegasus Construction Group Inc.", "phone": "727-480-4796"},
{"business_name": "Petco Grooming — Lady Lake", "phone": "352-430-7228"},
{"business_name": "PetSmart Grooming — The Villages", "phone": "352-750-3388"},
{"business_name": "Pierce Landscaping", "phone": "352-205-4543"},
{"business_name": "Premier Painting — Call Trina", "phone": "352-454-0062"},
{"business_name": "Pro Closet & Cabinetry", "phone": "352-694-9900"},
{"business_name": "Professional Organizer of The Villages — Margey", "phone": "772-707-6064"},
{"business_name": "Puddle Pool Services", "phone": "352-307-1898"},
{"business_name": "Q Services LLC", "phone": "352-638-0472"},
{"business_name": "Quality Moving Services", "phone": "352-633-8217"},
{"business_name": "Quality Moving Services LLC", "phone": "352-430-7393"},
{"business_name": "R&C Brady Services LLC", "phone": "352-461-0428"},
{"business_name": "RoadRunner Auto Transport — The Villages", "phone": "888-777-2123"},
{"business_name": "Ramsey's Pressure Washing", "phone": "352-220-9783"},
{"business_name": "Reliable Tasks Pros", "phone": "352-801-0288"},
{"business_name": "Residential Water Works LLC", "phone": "352-454-0303"},
{"business_name": "Ricardo's Painting Services LLC", "phone": "352-399-0671"},
{"business_name": "Roam Electric Inc.", "phone": "352-327-4434"},
{"business_name": "Roberto's Flooring", "phone": "352-552-1819"},
{"business_name": "Rock Raiders by Tip Top Trim", "phone": "352-268-4242"},
{"business_name": "Roof Guys", "phone": "239-390-1997"},
{"business_name": "S&W Roofing LLC", "phone": "352-319-6787"},
{"business_name": "Safe & Sound Roofing LLC", "phone": "352-942-9572"},
{"business_name": "San Juan Tile LLC", "phone": "352-399-2254"},
{"business_name": "Scott's Construction Unlimited LLC", "phone": "352-430-3885"},
{"business_name": "Selina Nails & Day Spa", "phone": "352-492-9526"},
{"business_name": "Senior Helpers — The Villages", "phone": "352-323-6232"},
{"business_name": "Sharie's Salon / Permanent Makeup by Sharie", "phone": "352-209-9048"},
{"business_name": "Shepherds Electric", "phone": "352-643-0340"},
{"business_name": "Signature Cart Detailing", "phone": "518-441-8782"},
{"business_name": "Skinner's Environmental & Home Improvements LLC", "phone": "352-430-1296"},
{"business_name": "Sod Lot — The Villages, The", "phone": "352-470-0326"},
{"business_name": "Sparkle Effect Mobile Detailing", "phone": "352-448-7707"},
{"business_name": "Speedy's RV & Boat Storage", "phone": "352-689-9699"},
{"business_name": "SprinkleRight Irrigation", "phone": "352-235-9829"},
{"business_name": "Stamped and Sealed Handyman LLC", "phone": "352-561-2020"},
{"business_name": "Summerfield Roofing", "phone": "352-421-3336"},
{"business_name": "Sunstate Power — The Villages", "phone": "352-561-3621"},
{"business_name": "Sunshine Awnings — Outdoor Cleaning", "phone": "352-787-1440"},
{"business_name": "TMT Beauty", "phone": "352-744-9902"},
{"business_name": "TNA Residential Services", "phone": "352-633-0311"},
{"business_name": "Tallman Lawn Services", "phone": "352-431-6326"},
{"business_name": "Taz Mania Landscaping LLC", "phone": "352-619-5772"},
{"business_name": "Thomas's Handyman Service", "phone": "352-234-4248"},
{"business_name": "Tire Guy — Andy Robinson, The", "phone": "352-434-2254"},
{"business_name": "Todd Casey Golf Cart Repairs", "phone": "352-399-7678"},
{"business_name": "Top Nails Salon", "phone": "352-750-2289"},
{"business_name": "Trev's Precision - Mobile Cart Service", "phone": "352-615-1027"},
{"business_name": "Tri-County Movers", "phone": "352-208-5288"},
{"business_name": "Two Men and a Truck — Ocala/The Villages", "phone": "352-347-3337"},
{"business_name": "U Save Door and Trim", "phone": "407-843-5606"},
{"business_name": "VIP Transportation Group", "phone": "352-353-1037"},
{"business_name": "Veterans Garage Door", "phone": "352-844-3925"},
{"business_name": "Village Airport Van", "phone": "352-241-2000"},
{"business_name": "Village Home Watch & Services, LLC", "phone": "352-456-8587"},
{"business_name": "Village Mover", "phone": "352-751-2750"},
{"business_name": "Village Pet Spa", "phone": "352-751-2877"},
{"business_name": "Village Tax & Accounting Solutions LLC", "phone": "352-259-7053"},
{"business_name": "Village Veterinarian", "phone": "352-751-2525"},
{"business_name": "Villages AV — Smart Home Installation", "phone": "352-388-1677"},
{"business_name": "Villages Computer Guy, The", "phone": "352-633-0287"},
{"business_name": "Villages Golf Cars, The", "phone": "352-750-4653"},
{"business_name": "Villages Home Watch LLC", "phone": "352-386-0444"},
{"business_name": "Villages Mobile Notary, The", "phone": "501-813-0123"},
{"business_name": "Villages Notary", "phone": "352-391-6600"},
{"business_name": "Villages Outdoor, The", "phone": "352-753-6290"},
{"business_name": "Villages Pet Sitters Network, The", "phone": "352-391-3016"},
{"business_name": "Villages Refrigerator & Appliance Repair", "phone": "352-348-3411"},
{"business_name": "VillagersHome WatchPlus", "phone": "518-423-9966"},
{"business_name": "Visiting Angels — The Villages", "phone": "352-643-0702"},
{"business_name": "Vivint Smart Home — The Villages", "phone": "305-894-0622"},
{"business_name": "Volthom Electric Inc.", "phone": "352-789-2454"},
{"business_name": "Warren's Landscaping & Property Maintenance", "phone": "352-603-9554"},
{"business_name": "Welcome Legacy Maintenance LLC", "phone": "352-512-1476"},
{"business_name": "West Orange Roofing", "phone": "352-770-8920"},
{"business_name": "Willie's Golf Cart Service & Repair", "phone": "352-255-1369"},
{"business_name": "Windows of Central Florida", "phone": "352-217-8151"},
{"business_name": "Woodies Flooring and More LLC", "phone": "352-560-1204"},
{"business_name": "Woof Gang Bakery & Grooming — The Villages", "phone": "352-633-0097"},
{"business_name": "iFixed EZ Florida", "phone": "352-461-2388"},
]

# Sort A-Z
providers_sorted = sorted(providers, key=lambda x: x['business_name'].lstrip('#').strip().upper())

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "V-HUB Providers"

# Colors
orange_red = "E8431A"
teal = "00BFA5"
white = "FFFFFF"
light_gray = "F5F5F5"
dark_gray = "333333"

# Header row
ws.append(["#", "Business Name", "Phone Number"])

# Style header
header_fill = PatternFill(start_color=orange_red, end_color=orange_red, fill_type="solid")
header_font = Font(bold=True, color=white, size=12)
header_border = Border(
    bottom=Side(style='medium', color=teal)
)
for col in range(1, 4):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = header_border

ws.row_dimensions[1].height = 24

# Data rows
thin_border = Border(
    bottom=Side(style='thin', color="DDDDDD")
)

for i, p in enumerate(providers_sorted, 1):
    phone = p['phone'] if p['phone'] else "—"
    ws.append([i, p['business_name'], phone])
    row = i + 1
    # Alternate row shading
    if i % 2 == 0:
        row_fill = PatternFill(start_color="FFF3F0", end_color="FFF3F0", fill_type="solid")
    else:
        row_fill = PatternFill(start_color=white, end_color=white, fill_type="solid")
    
    for col in range(1, 4):
        cell = ws.cell(row=row, column=col)
        cell.fill = row_fill
        cell.border = thin_border
        cell.font = Font(size=11, color=dark_gray)
        if col == 1:
            cell.alignment = Alignment(horizontal='center')
        elif col == 3:
            cell.alignment = Alignment(horizontal='center')

# Column widths
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 48
ws.column_dimensions['C'].width = 20

# Freeze top row
ws.freeze_panes = "A2"

# Add title row at very top
ws.insert_rows(1)
ws.merge_cells('A1:C1')
title_cell = ws['A1']
title_cell.value = "V-HUB Provider Directory — The Villages, FL"
title_cell.font = Font(bold=True, color=white, size=14)
title_cell.fill = PatternFill(start_color="1B3D6F", end_color="1B3D6F", fill_type="solid")
title_cell.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 30

# Save
wb.save('/app/VHub_Provider_Directory.xlsx')
print(f"Excel file created with {len(providers_sorted)} providers!")
